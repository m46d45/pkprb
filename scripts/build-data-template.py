#!/usr/bin/env python3
"""Build an Excel intake workbook: existing PKPRB data + empty rows to add."""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path("/workspace")
OUT = ROOT / "public" / "Template-Data-PKPRB.xlsx"

INK = "1F1A16"
PAPER = "F3EEE4"
TEAL = "3D7F92"
LINE = "D9D1C5"
EXIST = "F3EEE4"
NEW = "FFF8E8"
HDR = "1F1A16"
HINT = "6B6358"

thin = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)
fill_hdr = PatternFill("solid", fgColor=HDR)
fill_exist = PatternFill("solid", fgColor=EXIST)
fill_new = PatternFill("solid", fgColor=NEW)
fill_hint = PatternFill("solid", fgColor="EDE7DC")
font_hdr = Font(name="Calibri", bold=True, color="F3EEE4", size=10)
font_body = Font(name="Calibri", size=10, color=INK)
font_title = Font(name="Calibri", bold=True, size=16, color=INK)
font_h2 = Font(name="Calibri", bold=True, size=12, color=INK)
wrap = Alignment(wrap_text=True, vertical="center")

DISCIPLINE_LABELS = [
    "Teknik Sipil",
    "Arsitektur",
    "Planologi",
    "Teknik Geologi",
    "Teknik Lingkungan",
    "Teknik Kelautan",
    "Kebencanaan",
    "Multidisiplin",
]
DISC_FROM_ID = {
    "sipil": "Teknik Sipil",
    "arsitektur": "Arsitektur",
    "pwk": "Planologi",
    "geologi": "Teknik Geologi",
    "lingkungan": "Teknik Lingkungan",
    "kelautan": "Teknik Kelautan",
    "bencana": "Kebencanaan",
    "multidisiplin": "Multidisiplin",
}
STRATA = ["S1", "S2", "S3", "D4"]
PRODI_ACC = ["Internasional", "Unggul", "Baik Sekali", "Baik"]
IABEE = ["tidak", "provisional", "general"]
PT_ACC = ["Internasional", "Unggul", "Baik Sekali", "Baik"]
MATURITY = ["anchor", "pui", "standard"]
YA_TIDAK = ["ya", "tidak"]
HAZARDS = ["gempa", "tsunami", "banjir", "longsor", "likuefaksi", "gunungapi", "karhutla"]
STATUS = ["eksisting", "baru"]

PRODI_BLANK = 80
PT_BLANK = 40
CENTER_BLANK = 40
PKM_BLANK = 40


def extract_array(path: Path) -> list:
    text = path.read_text(encoding="utf-8")
    start = text.index("= [") + 2
    end = text.rindex("]") + 1
    blob = text[start:end]
    blob = re.sub(r",\s*]", "]", blob)
    return json.loads(blob)


def extract_universities(path: Path) -> list[dict]:
    text = path.read_text(encoding="utf-8")
    rows = []
    for m in re.finditer(
        r'"([^"]+)": \{ province: "([^"]+)", accreditation: "([^"]+)" \}',
        text,
    ):
        rows.append(
            {"university": m.group(1), "province": m.group(2), "accreditation": m.group(3)}
        )
    return rows


def style_header(ws: Worksheet, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(1, c)
        cell.fill = fill_hdr
        cell.font = font_hdr
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = thin
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(cols)}1"
    ws.sheet_properties.tabColor = TEAL


def write_row(ws: Worksheet, r: int, values: list, fill: PatternFill):
    for i, v in enumerate(values, 1):
        cell = ws.cell(r, i, v)
        cell.font = font_body
        cell.fill = fill
        cell.border = thin
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def widths(ws: Worksheet, ws_list: list[int]):
    for i, w in enumerate(ws_list, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def dv(ws: Worksheet, formula: str, cells: str):
    v = DataValidation(type="list", formula1=formula, allow_blank=True)
    v.error = "Pilih dari daftar"
    v.errorTitle = "Nilai tidak valid"
    v.prompt = "Pilih dari daftar"
    v.showErrorMessage = True
    v.showInputMessage = True
    ws.add_data_validation(v)
    v.add(cells)


def comment(ws: Worksheet, col: int, text: str):
    ws.cell(1, col).comment = Comment(text, "PKPRB")


def sheet_petunjuk(wb: Workbook, n_prodi: int, n_pt: int, n_center: int, n_prov: int):
    ws = wb.active
    ws.title = "00_Petunjuk"
    ws.sheet_properties.tabColor = "C45C48"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 92

    lines = [
        ("Template data PKPRB", font_title),
        ("Peta Keselarasan Pendidikan dan Risiko Bencana  ·  prototipe 20 Agustus 2026", Font(name="Calibri", italic=True, size=11, color=HINT)),
        ("", font_body),
        ("Untuk siapa", font_h2),
        ("Asisten / enumerator yang mencari data tambahan. Data yang sudah ada jangan dihapus. Baris berwarna krem = eksisting (boleh diperbaiki jika salah). Baris kuning muda = slot baru, isi di sini.", font_body),
        ("", font_body),
        ("Lembar", font_h2),
        (f"01_Prodi — {n_prodi} program studi yang sudah masuk peta. Tambah S1/S2/S3 Teknik Sipil, Arsitektur, Planologi, Geologi, Lingkungan, Kelautan, Kebencanaan, Multidisiplin. Prioritas: provinsi senjang (risiko tinggi, prodi sedikit) dan jenjang S2/S3 yang masih jarang.", font_body),
        (f"02_Perguruan_tinggi — {n_pt} institusi. Akreditasi PT (bukan prodi) dipakai spillover. Cari BAN-PT Perguruan Tinggi. Internasional hanya ITB, UI, UGM, ITS kecuali ada bukti akreditasi institusi internasional.", font_body),
        (f"03_Pusat_studi — {n_center} pusat. Tambah PUI, PSB, laboratorium kebencanaan. Isi bahaya yang relevan (kode, pisahkan koma).", font_body),
        ("04_Layanan_kepakaran — hampir kosong. Ini yang paling perlu. Layanan konsultasi, pendampingan BPBD, laboratorium uji, pelatihan SNI, klinik infrastruktur. Jangan menggandakan pusat studi kecuali ada layanan kepakaran yang jelas.", font_body),
        (f"05_Provinsi_risiko — {n_prov} provinsi. Penduduk dan skor risiko. Komposit dikalibrasi IRBI 2024 (Maluku 161,5; Malut 145,09; DKI 59,29). Skor per bahaya masih prototipe — ganti dengan angka resmi jika ada. Jangan ganti nama provinsi.", font_body),
        ("06_Daftar_kode — dropdown. Jangan diubah. Nama provinsi harus sama persis dengan daftar ini.", font_body),
        ("", font_body),
        ("Aturan isi", font_h2),
        ("1. Jangan hapus baris status = eksisting. Jika data salah, perbaiki selnya dan tulis di kolom catatan.", font_body),
        ("2. Baris baru: status = baru. Biarkan id kosong.", font_body),
        ("3. Nama provinsi wajib dari dropdown. Papua pemekaran: Papua, Papua Barat, Papua Selatan, Papua Tengah, Papua Pegunungan, Papua Barat Daya.", font_body),
        ("4. Satu baris = satu prodi (bukan satu universitas). S1 dan S2 sipil di kampus yang sama = dua baris.", font_body),
        ("5. IABEE hanya untuk prodi teknik. Jika IABEE general/provisional, akreditasi prodi boleh Internasional.", font_body),
        ("6. Wajib isi kolom sumber (URL BAN-PT, PDDikti, IABEE, laman LPPM). Tanpa sumber tidak masuk peta.", font_body),
        ("7. Kolom diisi_oleh dan tanggal_isi diisi nama + tanggal (YYYY-MM-DD).", font_body),
        ("", font_body),
        ("Sumber yang disarankan", font_h2),
        ("BAN-PT / LAM Teknik: https://www.banpt.or.id  dan laman LAM Teknik", font_body),
        ("PDDikti: https://pddikti.kemdiktisaintek.go.id", font_body),
        ("IABEE: https://iabee.or.id", font_body),
        ("IRBI / Inarisk BNPB: https://dibi.bnpb.go.id  ·  https://inarisk.bnpb.go.id", font_body),
        ("PUI-PT Kemdiktisaintek, laman LPPM masing-masing PT, direktori pusat studi kebencanaan.", font_body),
        ("", font_body),
        ("Yang diprioritaskan dicari dulu", font_h2),
        ("— S2/S3 Teknik Sipil dan Kebencanaan di luar Jawa.", font_body),
        ("— Planologi, Geologi, Lingkungan di Sulawesi, Maluku, Nusa Tenggara, Papua.", font_body),
        ("— Pusat studi dan layanan kepakaran di Aceh, Sumatera Barat, Sulawesi Tengah, NTT, Maluku.", font_body),
        ("— Akreditasi institusi BAN-PT yang resmi (Unggul / Baik Sekali / Baik), bukan tebakan dari prodi.", font_body),
        ("", font_body),
        ("Kirim file ini kembali utuh (jangan pecah per lembar) setelah terisi.", font_body),
        ("Peta: https://pkprb.vercel.app   Metodologi: https://pkprb.vercel.app/metodologi", font_body),
    ]
    for i, (text, font) in enumerate(lines, 1):
        cell = ws.cell(i, 1, text)
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 18 if len(text) < 80 else 36
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:2"
    ws.oddHeader.left.text = "PKPRB · template data"


def sheet_kode(wb: Workbook, provinces: list[str]) -> None:
    ws = wb.create_sheet("06_Daftar_kode")
    ws.sheet_properties.tabColor = LINE
    headers = [
        "status",
        "disiplin",
        "jenjang",
        "akreditasi_prodi",
        "iabee",
        "akreditasi_pt",
        "kematangan",
        "ya_tidak",
        "bahaya",
        "provinsi",
    ]
    style_header(ws, len(headers))
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    lists = {
        1: STATUS,
        2: DISCIPLINE_LABELS,
        3: STRATA,
        4: PRODI_ACC,
        5: IABEE,
        6: PT_ACC,
        7: MATURITY,
        8: YA_TIDAK,
        9: HAZARDS,
        10: provinces,
    }
    max_len = max(len(v) for v in lists.values())
    for r in range(2, max_len + 2):
        for c, vals in lists.items():
            if r - 2 < len(vals):
                cell = ws.cell(r, c, vals[r - 2])
                cell.font = font_body
                cell.border = thin
    widths(ws, [14, 22, 12, 20, 14, 18, 14, 12, 16, 32])
    ws.sheet_state = "visible"


def ref(col: str, n: int) -> str:
    return f"'06_Daftar_kode'!${col}$2:${col}${n}"


def sheet_prodi(wb: Workbook, programs: list[dict], provinces: list[str]):
    ws = wb.create_sheet("01_Prodi")
    headers = [
        "id",
        "status",
        "perguruan_tinggi",
        "nama_prodi",
        "disiplin",
        "jenjang",
        "akreditasi_prodi",
        "iabee",
        "kota",
        "provinsi",
        "sumber",
        "catatan",
        "diisi_oleh",
        "tanggal_isi",
    ]
    hints = [
        "Kosongkan jika baru",
        "eksisting / baru",
        "Nama resmi PT",
        "Contoh: Teknik Sipil",
        "Pilih dari daftar",
        "S1 S2 S3 D4",
        "Internasional jika IABEE",
        "tidak / provisional / general",
        "Kota/kabupaten kampus",
        "Wajib sama dengan daftar 38",
        "URL BAN-PT / PDDikti / IABEE",
        "Perbaikan atau keraguan",
        "Nama asisten",
        "YYYY-MM-DD",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
        comment(ws, i, hints[i - 1])
    style_header(ws, len(headers))

    def acc_label(p):
        if p.get("iabee") in ("general", "provisional"):
            return "Internasional"
        a = (p.get("accreditation") or "").lower()
        if "unggul" in a or a == "a":
            return "Unggul"
        if "baik sekali" in a or a == "b":
            return "Baik Sekali"
        return "Baik"

    r = 2
    for p in programs:
        iabee = p.get("iabee") or "none"
        write_row(
            ws,
            r,
            [
                p.get("id", ""),
                "eksisting",
                p.get("university", ""),
                p.get("program", ""),
                DISC_FROM_ID.get(p.get("discipline", ""), p.get("discipline", "")),
                p.get("strata", ""),
                acc_label(p),
                "tidak" if iabee == "none" else iabee,
                p.get("city", ""),
                p.get("province", ""),
                p.get("source", ""),
                "",
                "",
                "",
            ],
            fill_exist,
        )
        r += 1
    start_new = r
    for _ in range(PRODI_BLANK):
        write_row(ws, r, ["", "baru"] + [""] * 12, fill_new)
        r += 1
    last = r - 1
    widths(ws, [8, 12, 42, 42, 20, 10, 18, 14, 24, 28, 40, 28, 16, 14])
    nprov = 1 + len(provinces)
    dv(ws, ref("A", 3), f"B2:B{last}")
    dv(ws, ref("B", 9), f"E2:E{last}")
    dv(ws, ref("C", 5), f"F2:F{last}")
    dv(ws, ref("D", 5), f"G2:G{last}")
    dv(ws, ref("E", 4), f"H2:H{last}")
    dv(ws, ref("J", nprov), f"J2:J{last}")
    ws.auto_filter.ref = f"A1:N{last}"
    ws.freeze_panes = "C2"


def sheet_pt(wb: Workbook, unis: list[dict], provinces: list[str]):
    ws = wb.create_sheet("02_Perguruan_tinggi")
    headers = [
        "status",
        "perguruan_tinggi",
        "provinsi",
        "akreditasi_institusi",
        "sumber",
        "catatan",
        "diisi_oleh",
        "tanggal_isi",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    style_header(ws, len(headers))
    comment(ws, 4, "Akreditasi BAN-PT institusi, bukan prodi. Internasional hanya jika ada bukti institusional (bukan IABEE prodi).")
    acc_map = {
        "internasional": "Internasional",
        "unggul": "Unggul",
        "baik-sekali": "Baik Sekali",
        "baik": "Baik",
    }
    r = 2
    for u in unis:
        write_row(
            ws,
            r,
            [
                "eksisting",
                u["university"],
                u["province"],
                acc_map.get(u["accreditation"], u["accreditation"]),
                "",
                "prototipe — validasi BAN-PT institusi",
                "",
                "",
            ],
            fill_exist,
        )
        r += 1
    for _ in range(PT_BLANK):
        write_row(ws, r, ["baru"] + [""] * 7, fill_new)
        r += 1
    last = r - 1
    widths(ws, [12, 48, 28, 22, 40, 36, 16, 14])
    nprov = 1 + len(provinces)
    dv(ws, ref("A", 3), f"A2:A{last}")
    dv(ws, ref("J", nprov), f"C2:C{last}")
    dv(ws, ref("F", 5), f"D2:D{last}")
    ws.auto_filter.ref = f"A1:H{last}"


def sheet_centers(wb: Workbook, centers: list[dict], provinces: list[str]):
    ws = wb.create_sheet("03_Pusat_studi")
    headers = [
        "id",
        "status",
        "nama_pusat",
        "perguruan_tinggi",
        "provinsi",
        "bahaya",
        "kematangan",
        "jejaring_nasional",
        "ada_pkm",
        "url",
        "fokus",
        "sumber",
        "catatan",
        "diisi_oleh",
        "tanggal_isi",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    style_header(ws, len(headers))
    comment(ws, 6, "Kode bahaya dipisah koma, tanpa spasi berlebih. Contoh: gempa,tsunami,longsor. Daftar di 06_Daftar_kode kolom bahaya.")
    comment(ws, 7, "anchor = pusat rujukan nasional/internasional; pui = PUI-PT; standard = pusat biasa.")
    r = 2
    for c in centers:
        write_row(
            ws,
            r,
            [
                c.get("id", ""),
                "eksisting",
                c.get("name", ""),
                c.get("university", ""),
                c.get("province", ""),
                ",".join(c.get("hazards") or []),
                c.get("maturity", ""),
                "ya" if c.get("national") else "tidak",
                "ya" if c.get("pkm") else "tidak",
                c.get("url", ""),
                c.get("focus", ""),
                c.get("url", ""),
                "",
                "",
                "",
            ],
            fill_exist,
        )
        r += 1
    for _ in range(CENTER_BLANK):
        write_row(ws, r, ["", "baru"] + [""] * 13, fill_new)
        r += 1
    last = r - 1
    widths(ws, [8, 12, 48, 42, 28, 36, 14, 16, 12, 40, 44, 36, 24, 16, 14])
    nprov = 1 + len(provinces)
    dv(ws, ref("A", 3), f"B2:B{last}")
    dv(ws, ref("J", nprov), f"E2:E{last}")
    dv(ws, ref("G", 4), f"G2:G{last}")
    dv(ws, ref("H", 3), f"H2:H{last}")
    dv(ws, ref("H", 3), f"I2:I{last}")
    ws.auto_filter.ref = f"A1:O{last}"


def sheet_pkm(wb: Workbook, provinces: list[str]):
    ws = wb.create_sheet("04_Layanan_kepakaran")
    headers = [
        "status",
        "nama_layanan",
        "perguruan_tinggi",
        "provinsi",
        "jenis",
        "bahaya",
        "penerima",
        "url",
        "tahun_mulai",
        "sumber",
        "catatan",
        "diisi_oleh",
        "tanggal_isi",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    style_header(ws, len(headers))
    comment(ws, 5, "Jenis: konsultasi / pendampingan BPBD / laboratorium uji / pelatihan / klinik infrastruktur / lainnya.")
    comment(ws, 6, "Kode bahaya dipisah koma. Contoh: gempa,likuefaksi")
    comment(ws, 7, "Siapa yang dilayani: pemda, BPBD, PU, masyarakat, industri.")
    # one example row so the sheet isn't mysteriously empty
    write_row(
        ws,
        2,
        [
            "baru",
            "(contoh, hapus jika tidak valid) Klinik Rumah Aman Gempa",
            "Universitas Syiah Kuala",
            "Aceh",
            "pendampingan BPBD",
            "gempa,tsunami",
            "pemda / masyarakat",
            "",
            "",
            "isi URL",
            "Contoh format. Ganti atau hapus.",
            "",
            "",
        ],
        fill_new,
    )
    r = 3
    for _ in range(PKM_BLANK):
        write_row(ws, r, ["baru"] + [""] * 12, fill_new)
        r += 1
    last = r - 1
    widths(ws, [12, 42, 42, 28, 24, 28, 24, 36, 14, 36, 28, 16, 14])
    nprov = 1 + len(provinces)
    dv(ws, ref("A", 3), f"A2:A{last}")
    dv(ws, ref("J", nprov), f"D2:D{last}")
    ws.auto_filter.ref = f"A1:M{last}"


def sheet_provinsi(wb: Workbook, provinces: list[dict]):
    ws = wb.create_sheet("05_Provinsi_risiko")
    headers = [
        "status",
        "id",
        "provinsi",
        "pulau",
        "kode_bps",
        "penduduk",
        "risiko_komposit",
        "gempa",
        "tsunami",
        "banjir",
        "longsor",
        "likuefaksi",
        "gunungapi",
        "karhutla",
        "sumber_penduduk",
        "sumber_irbi",
        "catatan",
        "diisi_oleh",
        "tanggal_isi",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    style_header(ws, len(headers))
    comment(ws, 6, "Jangan ganti nama provinsi. Penduduk: angka jiwa, bukan ribuan.")
    comment(ws, 7, "Komposit: kalibrasi IRBI 2024. Skala warna peta 0–200.")
    comment(ws, 8, "Skor bahaya prototipe 0–100. Ganti dengan angka resmi jika tersedia.")
    r = 2
    for p in provinces:
        risk = p.get("risk") or {}
        write_row(
            ws,
            r,
            [
                "eksisting",
                p.get("id", ""),
                p.get("name", ""),
                p.get("island", ""),
                p.get("kode", ""),
                p.get("population", ""),
                risk.get("composite", ""),
                risk.get("gempa", ""),
                risk.get("tsunami", ""),
                risk.get("banjir", ""),
                risk.get("longsor", ""),
                risk.get("likuefaksi", ""),
                risk.get("gunungapi", ""),
                risk.get("karhutla", ""),
                "",
                "IRBI 2024 publik (komposit dikalibrasi)",
                "Skor per bahaya bersifat prototipe",
                "",
                "",
            ],
            fill_exist,
        )
        r += 1
    last = r - 1
    widths(ws, [12, 16, 28, 14, 12, 14, 16, 10, 12, 10, 10, 12, 12, 12, 22, 36, 36, 16, 14])
    dv(ws, ref("A", 3), f"A2:A{last}")
    ws.auto_filter.ref = f"A1:S{last}"
    note = ws.cell(last + 2, 1, "Jangan tambah baris provinsi. Indonesia di peta ini 38 provinsi. Perbaiki angka di baris yang ada.")
    note.font = Font(name="Calibri", italic=True, size=10, color=HINT)


def sheet_celah(wb: Workbook, programs: list[dict], provinces: list[dict], centers: list[dict]):
    ws = wb.create_sheet("07_Celah_prioritas")
    ws.sheet_properties.tabColor = "C45C48"
    headers = ["provinsi", "jumlah_prodi", "jumlah_pusat", "disiplin_yang_ada", "catatan_prioritas"]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    style_header(ws, len(headers))
    by_prov: dict[str, list] = {p["name"]: [] for p in provinces}
    for p in programs:
        by_prov.setdefault(p["province"], []).append(p)
    centers_by = {}
    for c in centers:
        centers_by.setdefault(c["province"], []).append(c)
    r = 2
    ranked = []
    for name, plist in by_prov.items():
        discs = sorted({DISC_FROM_ID.get(p["discipline"], p["discipline"]) for p in plist})
        ranked.append((len(plist), name, discs, len(centers_by.get(name, []))))
    ranked.sort()
    for nprog, name, discs, ncent in ranked:
        note = ""
        if nprog <= 2:
            note = "Prodi sangat sedikit — prioritas pencarian."
        elif ncent == 0:
            note = "Belum ada pusat studi teridentifikasi."
        missing = [d for d in DISCIPLINE_LABELS if d not in discs]
        if missing and nprog > 0:
            extra = "Belum ada: " + ", ".join(missing[:4])
            note = (note + " " + extra).strip()
        write_row(
            ws,
            r,
            [name, nprog, ncent, ", ".join(discs) if discs else "—", note],
            fill_exist if nprog > 2 else fill_new,
        )
        r += 1
    widths(ws, [28, 14, 14, 52, 56])
    intro = ws.cell(
        r + 1,
        1,
        "Pakai lembar ini sebagai antrean kerja. Provinsi kuning: isi dulu. Tidak perlu mengisi semua disiplin di setiap provinsi.",
    )
    intro.font = Font(name="Calibri", italic=True, size=10, color=HINT)


def build():
    programs = extract_array(ROOT / "src/data/programs.ts")
    centers = extract_array(ROOT / "src/data/centers.ts")
    provinces = extract_array(ROOT / "src/data/provinces.ts")
    unis = extract_universities(ROOT / "src/data/universities.ts")
    prov_names = [p["name"] for p in provinces]

    wb = Workbook()
    sheet_petunjuk(wb, len(programs), len(unis), len(centers), len(provinces))
    sheet_kode(wb, prov_names)
    sheet_prodi(wb, programs, prov_names)
    sheet_pt(wb, unis, prov_names)
    sheet_centers(wb, centers, prov_names)
    sheet_pkm(wb, prov_names)
    sheet_provinsi(wb, provinces)
    sheet_celah(wb, programs, provinces, centers)

    order = [
        "00_Petunjuk",
        "01_Prodi",
        "02_Perguruan_tinggi",
        "03_Pusat_studi",
        "04_Layanan_kepakaran",
        "05_Provinsi_risiko",
        "07_Celah_prioritas",
        "06_Daftar_kode",
    ]
    wb._sheets = [wb[n] for n in order]

    OUT.parent.mkdir(exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT} ({OUT.stat().st_size} bytes)")
    print("prodi", len(programs), "pt", len(unis), "centers", len(centers), "prov", len(provinces))


if __name__ == "__main__":
    build()
